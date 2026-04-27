#!/usr/bin/env python3

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path("/Users/mirror-admin/repos/swfi-terminal-live-backend-clean")
DOWNLOADS = Path("/Users/mirror-admin/Downloads")
SEED_JSON = ROOT / "data" / "seed" / "public_pensions_actually_allocate.json"
OUTPUT_JSON = ROOT / "data" / "seed" / "public_pensions_audit.json"


def resolve_latest_download(pattern: str, fallback: Path) -> Path:
    candidates = sorted(DOWNLOADS.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else fallback


PENSIONS_XLSX = resolve_latest_download(
    "public_pensions_actually_allocate*.xlsx",
    DOWNLOADS / "public_pensions_actually_allocate(1).xlsx",
)
PI300_XLSX = resolve_latest_download(
    "PI_300_*.xlsx",
    DOWNLOADS / "PI_300_2017,2024.xlsx",
)
PI300_PDF = resolve_latest_download(
    "PI-300*.pdf",
    DOWNLOADS / "PI-300-2017,2024.pdf",
)

MARKET_FALLBACKS = [
    "South Korea",
    "New Zealand",
    "Switzerland",
    "Netherlands",
    "Australia",
    "Singapore",
    "Malaysia",
    "Denmark",
    "Germany",
    "Finland",
    "Norway",
    "Canada",
    "Sweden",
    "Taiwan",
    "Mexico",
    "Colombia",
    "Chile",
    "France",
    "Kuwait",
    "Vietnam",
    "U.K.",
    "U.S.",
    "India",
    "China",
    "Japan",
    "Brazil",
    "Peru",
]

PENSION_ALIAS_TO_PI300 = {
    "washington state investment board": "Washington State Board",
    "florida state board of administration": "Florida State Board",
    "north carolina invesmtent authority": "North Carolina",
    "minnesota state board of investment": "Minnesota State Board",
    "new york city retirement systems": "New York City Retirement",
    "employees retirement system of texas": "Texas Employees",
    "new jersey division of investment": "New Jersey Div. of Invest.",
    "massachusetts pension reserves investment management board": "Massachusetts PRIM",
    "teacher retirement system of texas": "Texas Teachers",
    "state teachers retirement system of ohio": "Ohio State Teachers",
    "teachers retirement system of georgia": "Georgia Teachers",
    "regents of the university of california": "California University",
    "los angeles county employees retirement association": "Los Angeles County Empl.",
    "pennsylvania state employees retirement system": "Pennsylvania Employees",
    "pennsylvania public school employees retirement system": "Pennsylvania School Empl.",
}


def iso_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


def to_number(value: object) -> float:
    text = normalize_text(value).replace(",", "").replace("$", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def human_number(value: float) -> str:
    if value >= 1_000_000_000_000:
        return f"${value / 1_000_000_000_000:.1f}T"
    if value >= 1_000_000_000:
        return f"${value / 1_000_000_000:.1f}B"
    if value >= 1_000_000:
        return f"${value / 1_000_000:.1f}M"
    return f"${value:,.0f}"


def normalize_name_key(value: object) -> str:
    text = normalize_text(value).lower()
    replacements = {
        "&": " and ",
        "div.": "division",
        "div ": "division ",
        "empl.": "employees",
        "empl ": "employees ",
        "invesmtent": "investment",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    stopwords = {
        "and",
        "the",
        "of",
        "fund",
        "funds",
        "pension",
        "pensions",
        "retirement",
        "system",
        "plan",
        "trust",
        "authority",
        "board",
        "public",
        "state",
        "common",
    }
    tokens = [token for token in text.split() if token and token not in stopwords]
    return " ".join(tokens)


def load_pension_seed() -> list[dict[str, object]]:
    if SEED_JSON.exists():
        payload = json.loads(SEED_JSON.read_text(encoding="utf-8"))
        return [dict(item) for item in payload.get("items", []) if isinstance(item, dict)]

    wb = load_workbook(PENSIONS_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_index = 0
    for index, row in enumerate(rows):
        if row and len(row) >= 10 and normalize_text(row[1]).lower() == "name":
            header_index = index
            break
    headers = [normalize_text(cell) for cell in rows[header_index]]
    items = []
    for raw in rows[header_index + 1 :]:
        if not raw or not raw[1]:
            continue
        record = {headers[idx]: raw[idx] if idx < len(raw) else "" for idx in range(len(headers))}
        institution_name = normalize_text(record.get("name"))
        assets = to_number(record.get("assets"))
        items.append(
            {
                "rank": int(to_number(record.get("#"))),
                "institution_name": institution_name,
                "state": normalize_text(record.get("state")),
                "assets": assets,
                "assets_display": human_number(assets) if assets else "Not disclosed",
                "person_name": normalize_text(record.get("Name")),
                "person_email": normalize_text(record.get("Email")),
                "person_title": normalize_text(record.get("Title")),
                "updated_in_swfi": normalize_text(record.get("Updated in SWFI")),
                "coverage_status": "ready_now" if record.get("Email") and record.get("Title") and record.get("Updated in SWFI") else "needs_contacts",
            }
        )
    return items


def extract_pi300_markets() -> list[str]:
    wb = load_workbook(PI300_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    markets = []
    if values:
        headers = [normalize_text(cell) for cell in values[0]]
        market_idx = headers.index("Market") if "Market" in headers else 1
        for row in values[1:]:
            market = normalize_text(row[market_idx] if market_idx < len(row) else "")
            if market:
                markets.append(market)
    deduped = sorted(set(markets) | set(MARKET_FALLBACKS), key=len, reverse=True)
    return deduped


def parse_pi300_asset(value: str) -> float:
    cleaned = value.replace("$", "").strip()
    if cleaned.count(".") > 1 and "," not in cleaned:
        cleaned = cleaned.replace(".", "")
    cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned) * 1_000_000
    except ValueError:
        return 0.0


def parse_pi300_first_edition() -> list[dict[str, object]]:
    markets = extract_pi300_markets()
    text = "\n".join(page.extract_text() or "" for page in PdfReader(PI300_PDF).pages)
    lines = [normalize_text(line) for line in text.splitlines() if normalize_text(line)]
    rows: list[tuple[int, str]] = []
    seen_ranks: set[int] = set()
    for index, line in enumerate(lines):
        match = re.match(r"^(\d{1,3})\.\s+(.*)$", line)
        if not match:
            continue
        rank = int(match.group(1))
        if rank == 1 and seen_ranks and len(seen_ranks) >= 250:
            break
        body = match.group(2)
        cursor = index + 1
        while "$" not in body and cursor < len(lines) and cursor <= index + 3:
            if re.match(r"^\d{1,3}\.\s+", lines[cursor]):
                break
            body = f"{body} {lines[cursor]}"
            cursor += 1
        if "$" not in body or rank in seen_ranks:
            continue
        seen_ranks.add(rank)
        rows.append((rank, body))

    parsed: list[dict[str, object]] = []
    asset_pattern = re.compile(r"^(.*?)(\$[0-9,\.]+)(?:\s+\d+)?$")
    for rank, body in rows:
        match = asset_pattern.match(body)
        if not match:
            continue
        label = normalize_text(match.group(1))
        asset_text = match.group(2)
        market = ""
        fund = label
        for candidate in markets:
            if label.endswith(f" {candidate}"):
                fund = label[: -len(candidate)].strip()
                market = candidate
                break
        parsed.append(
            {
                "rank": rank,
                "fund": fund,
                "market": market,
                "total_assets": parse_pi300_asset(asset_text),
                "total_assets_display": asset_text,
                "name_key": normalize_name_key(fund),
            }
        )
    return parsed


def match_benchmark_record(name: str, benchmark: list[dict[str, object]]) -> dict[str, object] | None:
    query = normalize_name_key(name)
    if not query:
        return None
    alias = PENSION_ALIAS_TO_PI300.get(normalize_text(name).lower())
    if alias:
        alias_key = normalize_name_key(alias)
        for item in benchmark:
            if str(item.get("name_key") or "") == alias_key:
                return {
                    "rank": int(item.get("rank") or 0),
                    "fund": str(item.get("fund") or ""),
                    "market": str(item.get("market") or ""),
                    "total_assets": float(item.get("total_assets") or 0.0),
                    "total_assets_display": str(item.get("total_assets_display") or ""),
                    "match_score": 1.0,
                    "match_status": "verified",
                    "match_basis": "alias",
                }
    best_verified: dict[str, object] | None = None
    best_candidate: dict[str, object] | None = None
    best_score = 0.0
    query_tokens = query.split()
    for item in benchmark:
        candidate = str(item.get("name_key") or "")
        if not candidate:
            continue
        candidate_tokens = candidate.split()
        score = SequenceMatcher(None, query, candidate).ratio()
        same_first = bool(query_tokens and candidate_tokens and query_tokens[0] == candidate_tokens[0])
        if query == candidate:
            best_verified = {
                "rank": int(item.get("rank") or 0),
                "fund": str(item.get("fund") or ""),
                "market": str(item.get("market") or ""),
                "total_assets": float(item.get("total_assets") or 0.0),
                "total_assets_display": str(item.get("total_assets_display") or ""),
                "match_score": 1.0,
                "match_status": "verified",
                "match_basis": "normalized_exact",
            }
            break
        if not (score >= 0.84 or (same_first and score >= 0.72)):
            continue
        if score > best_score:
            best_candidate = {
                "rank": int(item.get("rank") or 0),
                "fund": str(item.get("fund") or ""),
                "market": str(item.get("market") or ""),
                "total_assets": float(item.get("total_assets") or 0.0),
                "total_assets_display": str(item.get("total_assets_display") or ""),
                "match_score": round(score, 4),
                "match_status": "candidate",
                "match_basis": "heuristic",
            }
            best_score = score
    return best_verified or best_candidate


def build_audit_payload() -> dict[str, object]:
    items = load_pension_seed()
    benchmark = [item for item in parse_pi300_first_edition() if item.get("market") == "U.S."]
    benchmark_matches = []
    benchmark_candidates = []
    unmatched = []
    critical_contact_gaps = []
    coverage_counter = Counter()
    state_counter = Counter()
    for item in items:
        state = normalize_text(item.get("state") or "")
        if state:
            state_counter[state] += 1
        coverage_status = str(item.get("coverage_status") or "needs_verification")
        coverage_counter[coverage_status] += 1
        assets = float(item.get("assets") or 0.0)
        record = {
            "rank": int(item.get("rank") or 0),
            "institution_name": str(item.get("institution_name") or ""),
            "state": state,
            "assets": assets,
            "assets_display": str(item.get("assets_display") or human_number(assets)),
            "coverage_status": coverage_status,
            "person_name": str(item.get("person_name") or ""),
            "person_email": str(item.get("person_email") or ""),
            "person_title": str(item.get("person_title") or ""),
        }
        match = match_benchmark_record(record["institution_name"], benchmark)
        if match:
            match_record = {
                **record,
                "pi300_rank": match["rank"],
                "pi300_fund": match["fund"],
                "pi300_market": match["market"],
                "pi300_total_assets": match["total_assets"],
                "pi300_total_assets_display": match["total_assets_display"],
                "match_score": match["match_score"],
                "match_status": match["match_status"],
                "match_basis": match["match_basis"],
            }
            if match["match_status"] == "verified":
                benchmark_matches.append(match_record)
            else:
                benchmark_candidates.append(match_record)
        else:
            unmatched.append(record)
        if coverage_status != "ready_now":
            critical_contact_gaps.append(record)

    benchmark_matches.sort(key=lambda item: (item["pi300_rank"], -item["assets"], item["institution_name"]))
    benchmark_candidates.sort(key=lambda item: (-item["assets"], item["pi300_rank"], item["institution_name"]))
    unmatched.sort(key=lambda item: (-item["assets"], item["institution_name"]))
    critical_contact_gaps.sort(key=lambda item: (-item["assets"], item["institution_name"]))
    top_100_overlap = sum(1 for item in benchmark_matches if int(item.get("pi300_rank") or 0) <= 100)
    top_50_overlap = sum(1 for item in benchmark_matches if int(item.get("pi300_rank") or 0) <= 50)
    benchmark_assets = sum(float(item.get("assets") or 0.0) for item in benchmark_matches)
    return {
        "schema_version": "swfi.public_pensions_audit.v1",
        "generated_at": iso_now(),
        "source_files": {
            "pensions_workbook": str(PENSIONS_XLSX),
            "pi300_pdf": str(PI300_PDF),
            "pi300_workbook": str(PI300_XLSX),
            "seed_json": str(SEED_JSON),
        },
        "summary": {
            "institutions": len(items),
            "ready_now": coverage_counter.get("ready_now", 0),
            "needs_contacts": coverage_counter.get("needs_contacts", 0),
            "needs_verification": coverage_counter.get("needs_verification", 0),
            "pi300_us_overlap": len(benchmark_matches),
            "pi300_review_candidates": len(benchmark_candidates),
            "pi300_top_100_overlap": top_100_overlap,
            "pi300_top_50_overlap": top_50_overlap,
            "critical_contact_gaps": len(critical_contact_gaps),
            "benchmark_assets": benchmark_assets,
            "benchmark_assets_display": human_number(benchmark_assets) if benchmark_assets else "Not disclosed",
            "top_states": [{"state": state, "count": count} for state, count in state_counter.most_common(12)],
        },
        "position": {
            "label": "Public pensions benchmark audit",
            "note": "Overlay the SWFI public-pensions target queue against the 2024 PI-300 U.S. cohort to expose where the highest-value allocator records still lack key-person coverage.",
        },
        "benchmark": {
            "label": "Thinking Ahead Institute / P&I 300",
            "year": 2024,
            "market_scope": "U.S. institutions only for direct overlap in this audit",
            "rows_parsed": len(benchmark),
        },
        "benchmark_matches": benchmark_matches[:75],
        "benchmark_review_candidates": benchmark_candidates[:50],
        "critical_contact_gaps": critical_contact_gaps[:50],
        "highest_asset_unmatched": unmatched[:40],
    }


def main() -> int:
    payload = build_audit_payload()
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(payload["summary"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
