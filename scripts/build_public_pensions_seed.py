#!/usr/bin/env python3

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def resolve_latest_download(pattern: str, fallback: Path) -> Path:
    candidates = sorted(Path("/Users/mirror-admin/Downloads").glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else fallback


DEFAULT_SOURCE = resolve_latest_download(
    "public_pensions_actually_allocate*.xlsx",
    Path("/Users/mirror-admin/Downloads/public_pensions_actually_allocate(1).xlsx"),
)
DEFAULT_OUTPUT = Path("/Users/mirror-admin/repos/swfi-terminal-live-backend-clean/data/seed/public_pensions_actually_allocate.json")


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


def profile_slug(value: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "-", normalize_text(value).lower()).strip("-")


def human_number(value: float) -> str:
    if value >= 1_000_000_000_000:
        return f"${value / 1_000_000_000_000:.1f}T"
    if value >= 1_000_000_000:
        return f"${value / 1_000_000_000:.1f}B"
    return f"${value:,.0f}"


def to_number(value: object) -> float:
    text = normalize_text(value).replace(",", "").replace("$", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def derive_status(institution_name: str, assets: float, email: str, title: str, updated: str) -> tuple[str, str, str]:
    if email and title and updated:
        return ("ready_now", "High", "Export-ready pension allocator record.")
    if institution_name and (assets or title or updated):
        return ("needs_contacts", "Medium", "Institution is present, but CIO or Executive Director coverage is incomplete.")
    return ("needs_verification", "Watch", "Institution needs key-person verification and SWFI refresh.")


def main() -> int:
    wb = load_workbook(DEFAULT_SOURCE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_index = 0
    for index, row in enumerate(rows):
        if row and len(row) >= 10 and normalize_text(row[1]).lower() == "name":
            header_index = index
            break
    headers = [normalize_text(cell) for cell in rows[header_index]]
    items = []
    for raw in rows[header_index + 1 :]:
        if not raw or not raw[1]:
            continue
        record = {headers[idx]: raw[idx] if idx < len(raw) else "" for idx in range(len(headers))}
        assets = to_number(record.get("assets"))
        email = normalize_text(record.get("Email"))
        title = normalize_text(record.get("Title"))
        updated = normalize_text(record.get("Updated in SWFI"))
        institution_name = normalize_text(record.get("name"))
        coverage_status, watch_priority, next_best_action = derive_status(institution_name, assets, email, title, updated)
        items.append(
            {
                "rank": int(record.get("#") or 0),
                "institution_name": institution_name,
                "slug": profile_slug(institution_name),
                "city": normalize_text(record.get("city")),
                "state": normalize_text(record.get("state")),
                "assets": assets,
                "assets_display": human_number(assets) if assets else "Not disclosed",
                "allocator_role_hint": normalize_text(record.get("CEO or CIO or Admnistrator")),
                "person_name": normalize_text(record.get("Name")),
                "person_email": email,
                "person_title": title,
                "updated_in_swfi": updated,
                "coverage_status": coverage_status,
                "export_readiness": "ready" if coverage_status == "ready_now" else "review",
                "watch_priority": watch_priority,
                "next_best_action": next_best_action,
                "source_refs": [
                    {
                        "label": "Public pensions allocation workbook",
                        "document_pointer": str(DEFAULT_SOURCE)
                    }
                ]
            }
        )
    items.sort(key=lambda item: (item["rank"], item["institution_name"]))
    payload = {
        "schema_version": "swfi.public_pensions_coverage.v1",
        "source_file": str(DEFAULT_SOURCE),
        "count": len(items),
        "items": items
    }
    DEFAULT_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    DEFAULT_OUTPUT.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
